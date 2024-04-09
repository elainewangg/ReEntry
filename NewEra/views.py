import ast
import json
import operator
import os
import pprint
import re
from datetime import date, datetime, timedelta
from dateutil.parser import isoparse
from decimal import Decimal
from pathlib import Path
from itertools import chain

import openpyxl
from django.contrib import messages
from django.contrib.auth import authenticate
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import IntegrityError, transaction
from django.db.models import Avg, Prefetch, Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.clickjacking import xframe_options_sameorigin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import pytz
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.core import mail
from twilio.rest import Client

# from NewEra.forms import (BiweeklyForm, CaseLoadUserForm, CreateNoteForm,
#                           CreateResourceForm, EditOrganizationForm,
#                           EditReferralNotesForm, EditSelfUserForm, EditStudentReferralNotesForm,
#                           EditUserForm, LoginForm, MeetingTrackerForm,
#                           RegistrationForm, ResourceFilter, RiskAssessmentForm,
#                           SelectDataTimeframe, StudentForm,
#                           StudentQuarterlyUpdateForm, StudentWeeklyUpdateForm,
#                           StudentWeeklyUpdateUploadFormset, TagForm, RoleSwitchForm)
from NewEra.forms import (CaseLoadUserForm, TempCaseLoadUserForm, CreateNoteForm,
                          CreateResourceForm, EditOrganizationForm,
                          EditReferralNotesForm, EditSelfUserForm,
                          EditUserForm, LoginForm, MeetingTrackerForm,
                          RegistrationForm, ResourceFilter,
                          SelectDataTimeframe, TagForm, RoleSwitchForm)
# from NewEra.models import (Biweekly, CaseLoadUser, MeetingTracker, Note,
#                            Organization, Referral, Resource, RiskAssessment,
#                            Student, StudentQuarterlyUpdate, StudentReferral,
#                            StudentWeeklyUpdate, Tag, User)
from NewEra.models import (CaseLoadUser, TempCaseLoadUser, MeetingTracker, Note,
                           Organization, Referral, Resource,
                           Tag, User)

# region CONSTANTS

###################################### 
#              CONSTANTS 
######################################

# Define the number of referrals or paginations shown on a page
RESOURCE_PAGINATION_COUNT = 30
REFERRAL_PAGINATION_COUNT = 20

# endregion

# region MISC ACTIONS

###################################### 
#              MISC ACTIONS  
######################################

# function to display home page
def home(request): 

    context = {}
    markReferralAsSeen(request)
    
    if 'confirmuser' in request.GET:
        confirm_user(request)
        context['confirmed'] = True
    

    return render(request, 'NewEra/home.html', context)

# function to login
def login(request):
    if request.user.is_authenticated:
        return redirect(reverse('Home'))

    context = {}

    if request.method == 'GET':
        context['form'] = LoginForm()
        return render(request, 'NewEra/login.html', context)

    form = LoginForm(request.POST)
    context['form'] = form

    if not form.is_valid():
        return render(request, 'NewEra/login.html', context)

    user = authenticate(username=form.cleaned_data['username'],
                            password=form.cleaned_data['password'])

    auth_login(request, user)

    messages.success(request, 'Logged in as {} {}.'.format(user.first_name, user.last_name))
    role = user.get_role()
    if not role:
        return render(request, 'NewEra/login.html', context)
    else:
        request.session['role'] = role

    if user.is_superuser:
        return redirect(reverse('Dashboard'))
    else:
        return redirect(reverse('Home'))

# Sends an email to the staff_user
def sendEmail(load_user):
    # Create email contents
    subject = 'Realistic ReEntry Sign Up from {}'.format(load_user.get_full_name())
    html_message = render_to_string('NewEra/sign_up_mailer.html',
                                    {'client': load_user})
    plain_message = strip_tags(html_message)
    from_email = settings.EMAIL_HOST_USER

    # Set recipient
    # to = settings.EMAIL_HOST_USER
    to = "aeli@andrew.cmu.edu"

    # Send the email
    mail.send_mail(subject, plain_message, from_email, [to], html_message=html_message, fail_silently=True)

def sendEmailConfirmation(load_case_user):
    try:
        subject = 'You have signed up for RealisticReEntry'
        html_message = render_to_string('NewEra/email_confirmation.html', 
                        {'tempid': load_case_user.id})
        plain_message = strip_tags(html_message)
        from_email = settings.EMAIL_HOST_USER
        to = load_case_user.email

        mail.send_mail(subject, plain_message, from_email, [to], html_message=html_message, fail_silently=False)
        return True
    except:
        return False

def sendSMSConfirmation(load_case_user):
        try:
            to = load_case_user.phone
            # Set the message intro string based on whether the referral is to someone on the case load or out of the system
            messageIntro = 'Thank you for signing up for ReEntry Services at RealisticReEntry!'

            # # Create the query string and the message body
            # queryString = '?key=' + referralTimeStamp
            # queryString = queryString.replace(' ', '%20')  # Make SMS links accessible
            # link = 'http://newera412.com/resources/' + str(r.id) + queryString
            # links = ['http://realisticreentry/']
            queryString = '?confirmuser='+ str(load_case_user.id)
            messageBody = 'Please click on this link to confirm your signup: http://127.0.0.1:8000' + queryString


            client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
            client.messages.create(from_=settings.TWILIO_PHONE_NUMBER, to=to, body=messageIntro + messageBody)
            return True
        except:
            return False

# Function to convert temp user to caseload user when they confirm signup 
def confirm_user(request):
    if 'confirmuser' not in request.GET:
        return 

    try:
        key_str = request.GET['confirmuser']
    except ValueError as e:
        # Log the error and return
        print(f"Error getting 'confirmuser' parameter: {str(e)}")
        return

    tempUsers = TempCaseLoadUser.objects.filter(id=key_str)

    if tempUsers.count() == 1:
        tempUser = tempUsers.first()
        new_case_load = CaseLoadUser.objects.create(first_name=tempUser.first_name, 
                                        last_name=tempUser.last_name,
                                        nickname=tempUser.nickname,
                                        email=tempUser.email, 
                                        phone=tempUser.phone, 
                                        neighborhood=tempUser.neighborhood,
                                        case_label=tempUser.case_label,
                                        is_active=tempUser.is_active,
                                        user=None,
                                        age=tempUser.age,
                                        zip_code=tempUser.zip_code,
                                        education=tempUser.education,
                                        is_vote_registered=tempUser.is_vote_registered)
        
        messages.success(request, '{} {} has successfully signed up.'.format(tempUser.first_name, tempUser.last_name))
        tempUser.delete()
        new_case_load.save()

    sendEmail(new_case_load)
    return redirect(reverse('Home'))

# function to sign up
def sign_up(request):
    context = {}

    if request.method == 'GET':
        context['form'] = TempCaseLoadUserForm()
        return render(request, 'NewEra/sign_up.html', context)
    if request.method == 'POST':
        staff_user = User.objects.filter(is_superuser=True).first()
        load_user = TempCaseLoadUser(user=staff_user)
        form = TempCaseLoadUserForm(request.POST, instance=load_user)
        if not form.is_valid():
            context['form'] = form 
            return render(request, 'NewEra/sign_up.html', context)
        form.save()
        load_user.save()
        confirmationSent = False
        if load_user.email:
            confirmationSent = sendEmailConfirmation(load_user)
        elif load_user.phone:
            confirmationSent = sendSMSConfirmation(load_user)
        if confirmationSent:
            messages.success(request, 'A confirmation message has been sent to you!')
        else:
            messages.error(request, 'A confirmation message was not able to be sent. Please provide a different email/phone number!')


    context['form'] = TempCaseLoadUserForm()
    return redirect(reverse('Home'))

def switch_role(request):
    if (request.method == 'POST'):
        form = RoleSwitchForm(request.POST)
        role = request.POST['role']

        request.session['role'] = role

        request.session.modified = True
    else:
        form = RoleSwitchForm()

    return redirect(reverse('Home'))

# function to logout
@login_required
def logout(request):
    auth_logout(request)
    messages.success(request, 'Successfully logged out.')
    return redirect(reverse('Login'))

# function to display about page
def about_us(request):
    return render(request, 'NewEra/about_us.html')

def programs(request):
    return render(request, 'NewEra/yarp.html')

def yarp(request):
    return render(request, 'NewEra/yarp.html')

def obb(request):
    return render(request, 'NewEra/obb.html')

def partners(request):
    return render(request, 'NewEra/partners.html')

# Function to update the referral given a GET request (to a resource) with a querystring timestamp
def markReferralAsSeen(request):

    if 'key' not in request.GET:
        return 

    try:
        key_str = request.GET['key'].replace(' 00:00', '')  # Remove ' 00:00' from the string
        keyDate = isoparse(key_str)
        keyDate = keyDate.replace(tzinfo=pytz.utc)  # Make the datetime object timezone-aware (assuming UTC)
    except ValueError as e:
        # Log the error and return
        print(f"Error parsing 'key' parameter: {str(e)}")
        return

    referrals = Referral.objects.filter(referral_date=keyDate)

    if referrals.count() == 1:
        referral = referrals.first()
        referral.date_accessed = datetime.now()
        referral.save()


# Function to check visitor cookie, and see if they accessed the resource
def isUniqueVisit(request, response, id): 
    siteStaff = request.COOKIES.get('siteStaff', '')

    if request.user.is_authenticated or siteStaff == 'true':
        response.set_cookie('siteStaff', 'true')
        return False 

    visitedResources = request.COOKIES.get('visitedResources', '').split(';')

    if visitedResources == ['']:
        response.set_cookie('visitedResources', str(id))
        return True 
    elif str(id) in visitedResources:
        return False
    else: 
        val = ';'.join(visitedResources)
        val = val + ';' + str(id)
        response.set_cookie('visitedResources', val)
        return True 
    
    return False 

# endregion

# region RESOURCES

###################################### 
#              RESOURCES  
######################################

# function to display all resources
def resources(request):
    all_resources = Resource.objects.all()
    context = { 'filter': ResourceFilter(request.GET, queryset=all_resources) }
    # context['employment_filter'] =  ResourceEmploymentFilter(request.GET, queryset=all_resources)
    # context = { 'employment_filter': ResourceEmploymentFilter(request.GET, queryset=all_resources) }
    # context['housing_filter'] = ResourceHousingFilter(request.GET, queryset=all_resources)
    # context['support_filter'] = ResourceSupportFilter(request.GET, queryset=all_resources)

    if request.method == 'GET':
        # SEARCH QUERY
        query = request.GET.get('query')

        # Set context value based on if the filter is set on the resources page
        if query:
            context['active_resources'] = context['filter'].qs.filter( Q(is_active=True) & (Q(name__icontains=query) | Q(description__icontains=query)) )
            context['inactive_resources'] = context['filter'].qs.filter( Q(is_active=False) & (Q(name__icontains=query) | Q(description__icontains=query)) )
        else: 
            context['active_resources'] = context['filter'].qs.filter(is_active=True)
            context['inactive_resources'] = context['filter'].qs.filter(is_active=False)

        # FILTER QUERY - build a query param for use in pagination links
        filterParams = request.GET.getlist('tags', '')
        filterQueryString = ''

        if filterParams: 
            for id in filterParams:
                filterQueryString += '&tags='
                filterQueryString += id

        context['filterQuery'] = filterQueryString

        # PAGINATION
        active_page = request.GET.get('a_page', 1)
        inactive_page = request.GET.get('i_page', 1)
        paginator = Paginator(context['active_resources'], RESOURCE_PAGINATION_COUNT)
        inactive_paginator = Paginator(context['inactive_resources'], RESOURCE_PAGINATION_COUNT)
        
        # Render the user's selected page for active resources
        try:
            activeResources = paginator.page(active_page)
        except PageNotAnInteger:
            activeResources = paginator.page(1)
        except EmptyPage:
            activeResources = paginator.page(paginator.num_pages)

        # Render the user's selected page for inactive resources
        try:
            inactiveResources = inactive_paginator.page(inactive_page)
        except PageNotAnInteger:
            inactiveResources = inactive_paginator.page(1)
        except EmptyPage:
            inactiveResources = inactive_paginator.page(inactive_paginator.num_pages)

        context['active_resources'] = activeResources
        context['inactive_resources'] = inactiveResources

    return render(request, 'NewEra/resources.html', context)

# function to display a certain resource detail
def get_resource(request, id):
    resource = get_object_or_404(Resource, id=id)
    context = { 'resource': resource, 'tags': resource.tags.all() }

    # provide a YouTube ID if this resource is a video
    if resource.resource_type == "video":
        pattern = r"(.*?)(^|\/|v=)([a-z0-9_-]{11})(.*)?"
        match = re.match(pattern, resource.url, re.IGNORECASE)
        if match and match.group(3):
            context['youtube_id'] = match.group(3)
    
    response = render(request, 'NewEra/get_resource.html', context)

    # Update the resource clicks
    if isUniqueVisit(request, response, id):
        resource.clicks = resource.clicks + 1
        resource.save()

    markReferralAsSeen(request)
    # markStudentReferralAsSeen(request)
    return response

# get resource attachment when uploading a new resource
@xframe_options_sameorigin
def get_resource_attachment(request, id): 
    resource = get_object_or_404(Resource, id=id)

    if not resource.image:
        raise Http404

    response = HttpResponse(resource.attachment, content_type=resource.attachment_content_type)
    response['Content-Disposition'] = 'filename="' + resource.attachment_content_type +'"'

    return response

# Deletes the given image if it exists
@login_required
def deleteAttachment(request, oldAttachment):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ATT_ROOT = os.path.join(BASE_DIR, 'NewEra/user_uploads/' + oldAttachment.name)
    os.remove(ATT_ROOT)

# ***** Note about images *****
# They are uploaded to the system as type .JPEG or .PNG etc.
# And then saved as type django.FileField() 
# *****************************
def get_resource_image(request, id): 
    resource = get_object_or_404(Resource, id=id)

    if not resource.image:
        raise Http404

    return HttpResponse(resource.image, content_type=resource.content_type)

# Deletes the given image if it exists
@login_required
def deleteImage(request, oldImage):
    if oldImage: 
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        IMAGE_ROOT = os.path.join(BASE_DIR, 'NewEra/user_uploads/' + oldImage.name)
        os.remove(IMAGE_ROOT)

# create new resources
@login_required
def create_resource(request):
    if not request.user.is_superuser:
        raise Http404

    context = {}
    form = CreateResourceForm()

    if request.method == 'POST':
        resource = Resource()
        form = CreateResourceForm(request.POST, request.FILES, instance=resource)
        
        if form.is_valid():
            # Update the Resource content_type MANUALLY 
            pic = form.cleaned_data['image']
            if pic and pic != '':
                resource.content_type = form.cleaned_data['image'].content_type

            att = form.cleaned_data['attachment']
            if att and att != '':
                resource.attachment_content_type = form.cleaned_data['attachment'].content_type

            form.save()
            resource.save()

            messages.success(request, 'Resource successfully created!')

            return redirect('Resources')
    else:
        form = CreateResourceForm(initial={'is_active': True})

    context['action'] = 'Create'
    context['form'] = form

    return render(request, 'NewEra/edit_resource.html', context)

# edit resource
@login_required
def edit_resource(request, id):
    if not request.user.is_superuser:
        raise Http404
    resource = get_object_or_404(Resource, id=id)
    oldImage = resource.image
    oldAttachment = resource.attachment

    if request.method == "POST":
        form = CreateResourceForm(request.POST, request.FILES, instance=resource)
    
        if form.is_valid():

            pic = form.cleaned_data['image']
            if pic and pic != '':
                
                # Update content type & remove old image
                try: 
                    # THE FOLLOWING LINE MAY THROW:
                    # After being set, the resource image is saved as a FileField type (and not an Image)
                    # As a result, it will not have a content_type attribute the way image files will
                    resource.content_type = form.cleaned_data['image'].content_type
                    deleteImage(request, oldImage)
                except: 
                    pass


                pic = form.cleaned_data['attachment']
            if pic and pic != '':
                
                # Update content type & remove old attachment
                try: 
                    resource.attachment_content_type = form.cleaned_data['attachment'].content_type
                    deleteAttachment(request, oldAttachment)
                except: 
                    pass

            form.save()
            resource.save()

            messages.success(request, '{} successfully edited.'.format(resource.name))
            return redirect('Show Resource', id=resource.id)
    else:
        form = CreateResourceForm(instance=resource)
    return render(request, 'NewEra/edit_resource.html', {'form': form, 'resource': resource, 'action': 'Edit'})

@login_required
def delete_resource(request, id):
    if not request.user.is_superuser:
        raise Http404

    resource = get_object_or_404(Resource, id=id)

    if request.method == 'POST':
        # Delete the resource (and its image) only if it has never been referred
        if (resource.referrals.count() == 0):
            deleteImage(request, resource.image)
            resource.delete()
            messages.success(request, '{} successfully deleted.'.format(resource.name))
            return redirect('Resources')
        # Otherwise, deactivate the resource
        else:
            resource.is_active = False
            resource.save()
            messages.success(request, '{} was made inactive.'.format(resource.name))
            return redirect('Show Resource', id=resource.id)
    return render(request, 'NewEra/delete_resource.html', {'resource': resource})

@login_required
def resetViews(request):
    if request.method == 'POST':
        Resource.objects.all().update(clicks=0)
        messages.success(request, 'Reset all resource views')
        return redirect(reverse('Dashboard'))
    return render(request, 'NewEra/reset_view_counts.html', {})

# endregion

# region REFERRALS

###################################### 
#              REFERRALS  
######################################

@login_required
def create_referral(request):
    resources = request.GET.get('resources', None)	
    recipients = []
    nameInputs = []
    referrals = []
    
    if request.user.is_superuser: 
        recipients = CaseLoadUser.objects.filter(is_active=True).all()
    elif request.user.is_superuser or request.user.is_reentry_coordinator or request.user.is_community_outreach_worker or request.user.is_service_provider or request.user.is_resource_coordinator:
        recipients = CaseLoadUser.objects.filter(is_active=True).filter(user=request.user)

    if request.method == 'GET' and resources:
        resources = [digit.strip() for digit in ast.literal_eval(resources)] # Safely parse array
        resources = [ get_object_or_404(Resource, id=resourceId) for resourceId in resources ]

        return render(request, 'NewEra/create_referral.html', {'resources': resources, 'recipients': recipients })

    elif request.method == 'POST': 
        phoneInput = ''.join(digit for digit in request.POST.get('phone', '') if digit.isdigit())
        
        # Referral to people on a case load (in the system)
        if 'resources[]' in request.POST and 'user_ids[]' in request.POST and 'notes' in request.POST: 
            caseload_users = [get_object_or_404(CaseLoadUser, id=num) for num in request.POST.getlist('user_ids[]')]
            resources = [get_object_or_404(Resource, id=num) for num in request.POST.getlist('resources[]')]

            for caseload_user in caseload_users:
                # Change the message introduction depending on whether the case load user has a nickname or not
                if caseload_user.nickname:
                    nameInputs.append(caseload_user.nickname)
                else:
                    nameInputs.append(caseload_user.first_name)

                referral = Referral(email=caseload_user.email, phone=caseload_user.phone, notes=request.POST['notes'], user=request.user, caseUser=caseload_user)
                referrals.append(referral)

        # Out of system referral
        elif 'resources[]' in request.POST and 'phone' in request.POST and 'email' in request.POST and 'notes' in request.POST and (len(phoneInput) == 10 or len(phoneInput) == 0): 
            resources = [get_object_or_404(Resource, id=num) for num in request.POST.getlist('resources[]')]
            referral = Referral(email=request.POST['email'], phone=phoneInput, notes=request.POST['notes'], user=request.user)
            referrals.append(referral)
            nameInput = request.POST['name']
            nameInputs.append(nameInput)

        else: 
            messages.error(request, 'Please fill out all fields.')
            return render(request, 'NewEra/create_referral.html', {'resources': resources, 'recipients': recipients })

        # add resource(s) to referral(s)
        for ref in referrals:
            ref.save()
            for r in resources: 
                ref.resource_set.add(r)
        
        # Send the referral(s) via email and SMS
        for i, name in enumerate(nameInputs):
            referralTimeStamp = str(referrals[i].referral_date)
            referrals[i].sendEmail(referralTimeStamp, name)
            referrals[i].sendSMS(referralTimeStamp, name)

    messages.success(request, 'Successfully created a new referral.')

    return redirect(reverse('Resources'))


def load_referrals(request, referrals, template_name):
    page = request.GET.get('page', 1)
    paginator = Paginator(referrals, REFERRAL_PAGINATION_COUNT)

    try:
        referrals = paginator.page(page)
    except PageNotAnInteger:
        referrals = paginator.page(1)
    except EmptyPage:
        referrals = paginator.page(paginator.num_pages)

    context = {'referrals': referrals}
    return render(request, template_name, context)

@login_required
def referrals(request):
    # Show admins all referrals, and SOWs only their referrals
    if (request.user.is_superuser):
        referrals = Referral.objects.all().order_by('-referral_date')
    elif (request.user.is_supervisor):
        referrals = Referral.objects.all().filter(user__in=User.objects.filter(organization=request.user.organization)).order_by('-referral_date')
    else:
        referrals = Referral.objects.all().filter(user=request.user).order_by('-referral_date')

    page = request.GET.get('page', 1)

    generalPaginator = Paginator(referrals, REFERRAL_PAGINATION_COUNT)

    try:
        referrals = generalPaginator.page(page)
    except PageNotAnInteger:
        referrals = generalPaginator.page(1)
    except EmptyPage:
        referrals = generalPaginator.page(generalPaginator.num_pages)

    context = { 'referrals': referrals }

    return render(request, 'NewEra/referrals.html', context)


@login_required
def get_referral(request, id):
    referral = get_object_or_404(Referral, id=id)
    if referral.user != request.user and not(request.user.is_superuser) and not (request.user.is_supervisor and request.user.organization == referral.user.organization ):
        raise Http404
    context = { 'referral': referral, 'resources': Resource.objects.all().filter(referrals=referral) }
    return render(request, 'NewEra/get_referral.html', context)

def edit_referral_notes(request, id):
    referral = get_object_or_404(Referral, id=id)
    if referral.user != request.user and not(request.user.is_superuser):
        raise Http404
    if request.method == "POST":
        form = EditReferralNotesForm(request.POST, instance=referral)
    
        if form.is_valid():
            form.save()
            referral.save()

            messages.success(request, 'Referral notes updated.')
            return redirect('Show Referral', id=referral.id)
    else:
        form = EditReferralNotesForm(instance=referral)
    return render(request, 'NewEra/edit_referral_notes.html', {'form': form, 'referral': referral, 'action': 'Edit'})

# region CASE LOAD
###################################### 
#              CASE LOAD  
######################################

@login_required
def case_load(request):
    users = []
    context = {} 

    # Set users to show differently based on the current user logged in
    if request.user.is_superuser: 
        users = CaseLoadUser.objects.filter(user__isnull=False)
        context['unassigned_caseload_users'] = CaseLoadUser.objects.filter(user=None)
        # Changed to account for inactive users
        context['staff'] = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    elif request.user.is_supervisor:
        users = CaseLoadUser.objects.filter(user__in=User.objects.filter(organization=request.user.organization)).order_by('first_name', 'last_name')
    elif request.user.is_reentry_coordinator or request.user.is_community_outreach_worker or request.user.is_service_provider or request.user.is_resource_coordinator:
        users = CaseLoadUser.objects.filter(user=request.user).order_by('first_name', 'last_name')
        print("HOLA")
    else:  
        raise Http404

    if request.method == 'POST' and 'staff_id' in request.POST:
        staff_user = get_object_or_404(User, id=request.POST['staff_id'])
        load_user = CaseLoadUser(user=staff_user)
        form = CaseLoadUserForm(request.POST, instance=load_user)
        
        if not form.is_valid():
            context['form'] = form 
            context['caseload_users'] = users
            context['modalStatus'] = 'show'
            return render(request, 'NewEra/case_load.html', context)
        form.save()
        load_user.user = staff_user
        load_user.save()
        messages.success(request, 'Successfully added {} {} to the CaseLoad.'.format(load_user.first_name, load_user.last_name))

    context['caseload_users'] = users
    if request.user.is_superuser:  
        context['form'] = CaseLoadUserForm()
    else:
        context['form'] = CaseLoadUserForm(instance=request.user)
    return render(request, 'NewEra/case_load.html', context)

@login_required
def get_case_load_user(request, id):
    case_load_user = get_object_or_404(CaseLoadUser, id=id)
    if case_load_user.user != request.user and not(request.user.is_superuser) and not (request.user.is_supervisor and request.user.organization == case_load_user.user.organization ):
        raise Http404
    notes = Note.objects.filter(case=case_load_user).order_by("-date")
    context = { 'case_load_user': case_load_user, 'notes': notes}

    return render(request, 'NewEra/get_case_load_user.html', context)

@login_required
def edit_case_load_user(request, id):
    case_load_user = get_object_or_404(CaseLoadUser, id=id)
    if case_load_user.user != request.user and not(request.user.is_superuser):
        raise Http404
    if request.method == "POST":
        form = CaseLoadUserForm(request.POST, instance=case_load_user)

        if form.is_valid():

            form.save()
            case_load_user.save()

            messages.success(request, '{} successfully edited.'.format(case_load_user.get_full_name()))
            return redirect('Show Case Load User', id=case_load_user.id)
    else:
        if request.user.is_superuser:  
            form = CaseLoadUserForm()
        else:
            form = CaseLoadUserForm(instance=request.user)
    return render(request, 'NewEra/edit_case_load_user.html', {'form': form, 'case_load_user': case_load_user, 'action': 'Edit'})

@login_required
def delete_case_load_user(request, id):
    case_load_user = get_object_or_404(CaseLoadUser, id=id)
    if case_load_user.user != request.user and not(request.user.is_superuser):
        raise Http404
    if request.method == 'POST':
        # Delete the case load user only if they have never been referred
        if (case_load_user.get_referrals().count() == 0):
            notes = Note.objects.filter(case=case_load_user)
            for n in notes:
                n.delete()
            case_load_user.delete()
            messages.success(request, '{} successfully deleted.'.format(case_load_user.get_full_name()))
            return redirect('Case Load')
        # Otherwise, delete the case load user
        else:
            case_load_user.is_active = False
            case_load_user.save()
            messages.success(request, '{} was made inactive.'.format(case_load_user.get_full_name()))
            return redirect('Show Case Load User', id=case_load_user.id)
    return render(request, 'NewEra/delete_case_load_user.html', {'case_load_user': case_load_user})

@login_required
def create_note(request, id):
    case_load_user = get_object_or_404(CaseLoadUser, id=id)
    if case_load_user.user != request.user and not(request.user.is_superuser):
        raise Http404
    context = {}
    form = CreateNoteForm()
    context['case_load_user'] = case_load_user 
    context['form'] = form
    context['action'] = 'Create'

    if request.method == 'POST':
        note = Note(case=case_load_user)
        form = CreateNoteForm(request.POST, instance=note)
        
        if form.is_valid():
            note.notes=form.cleaned_data['notes']
            note.activity_type=form.cleaned_data['activity_type']
            note.hours = form.cleaned_data['hours']
            note.date = form.cleaned_data['date']
            form.save()
            note.save()

            messages.success(request, 'Note successfully created!')

            return redirect('Show Case Load User', id=case_load_user.id)
    else:
        note = CreateNoteForm()

    return render(request, 'NewEra/edit_note.html', context)

@login_required
def edit_note(request, id):
    note = Note.objects.filter(id=id).first()
    
    case_load_user = note.case
    if case_load_user.user != request.user and not(request.user.is_superuser):
        raise Http404
    context = {}
    context['case_load_user'] = case_load_user 
    context['action'] = 'Edit'

    if request.method == "POST":
        form = CreateNoteForm(request.POST, instance=note)
        context['form'] = form

        if form.is_valid():
            form.save()
            note.save()

            messages.success(request, 'Note successfully edited!')
            return redirect('Show Case Load User', id=case_load_user.id)
            
    else:
        form = CreateNoteForm(instance=note)
        context['form'] = form
    return render(request, 'NewEra/edit_note.html', context)

@login_required
def delete_note(request, id):    
    note = Note.objects.filter(id=id).first()
    cid = note.case.id
    if note.case.user != request.user and not(request.user.is_superuser):
        raise Http404
    if request.method == 'POST':
        note.delete()
        messages.success(request, 'Note successfully deleted.')
        return redirect('Show Case Load User', id=cid)

    return render(request, 'NewEra/delete_note.html', {'note': note})

# endregion

# region MAPS
# def get_maps(request):
#     if not request.user.is_superuser:
#         raise Http404
    
#     # read json file of neighborhoods
#     with open('NewEra/static/NewEra/pgh_neighborhoods.geojson') as file:
#         neighborhoods = json.load(file)

#     # read json file of second neighborhoods
#     with open('NewEra/static/NewEra/Allegheny_County.geojson') as file:
#         second_neighborhoods = json.load(file)

#     neighborhoods['features'].extend(second_neighborhoods['features'])

#     # # Filter out entries with name "PITTSBURGH"
#     neighborhoods['features'] = [neighborhood for neighborhood in neighborhoods['features'] if not is_pittsburgh(neighborhood)]

#     for neighborhood in neighborhoods['features'] :
#         # Extract the neighborhood name from the appropriate field
#         if 'hood' in neighborhood['properties']:
#             name = neighborhood['properties']['hood']
#         elif 'NAME' in neighborhood['properties']:
#             name = neighborhood['properties']['LABEL']
#         else:
#             # Handle the case when the name field is not found in the properties
#             name = None  # Update this with appropriate handling
        
#         if name is not None:
#             # get the risk assessment
#             # risks = RiskAssessment.objects.filter(neighborhood=name).order_by('-risk_level')		
#             # risk = 0
#             # if len(risks) > 0 :
#             #     risk = risks[0].risk_level
#             # neighborhood['properties']['x_risk'] = risk

#             # get the active cases
#             active_cases = CaseLoadUser.objects.filter(neighborhood=name, is_active=True)
#             neighborhood['properties']['x_active_cases'] = len(active_cases)

#             # get the number of referrals
#             referrals = Referral.objects.filter(referral_date__gte=timezone.now()-timedelta(days=30), caseUser__neighborhood=name)
#             neighborhood['properties']['x_referrals'] = len(referrals)

#     output = json.dumps(neighborhoods)

#     context = { 'output': output }
    
#     response = render(request, 'NewEra/maps.html', context)

#     return response

# def is_pittsburgh(neighborhood):
#     if 'hood' in neighborhood['properties']:
#         name = neighborhood['properties']['hood']
#     elif 'NAME' in neighborhood['properties']:
#         name = neighborhood['properties']['NAME']
#     else:
#         name = None

#     return name == "PITTSBURGH"

# endregion

# region USER MANAGEMENT

###################################### 
#              USER MANAGEMENT  
######################################

@login_required
def dashboard(request): 
    if not request.user.is_superuser:
        raise Http404

    admins = User.objects.filter(is_superuser=True)
    reentry_coordinators = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_reentry_coordinator=True)
    community_outreach_workers = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_community_outreach_worker=True)
    service_providers = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_service_provider=True)
    resource_coordinators = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_resource_coordinator=True)
    supervisors = User.objects.filter(is_superuser=False).filter(is_supervisor=True)
    orgs = Organization.objects.all().order_by('name')
    context = {
        'admins': admins, 
        'reentry_coordinators': reentry_coordinators, 
        'community_outreach_workers': community_outreach_workers, 
        'service_providers': service_providers, 
        'resource_coordinators': resource_coordinators, 
        'orgs': orgs, 
        'supervisors': supervisors, 
        'form': RegistrationForm(), 
        'user' : request.user
    }

    if request.method == 'POST':
        if 'username' in request.POST:
            form = RegistrationForm(request.POST)
            context['form'] = form

            if not form.is_valid():
                context['modalStatus'] = 'show'
                return render(request, 'NewEra/dashboard.html', context)

            user = User.objects.create_user(username=form.cleaned_data['username'],
                                            password=form.cleaned_data['password'],
                                            email=form.cleaned_data['email'],
                                            phone=form.cleaned_data['phone'],
                                            first_name=form.cleaned_data['first_name'],
                                            last_name=form.cleaned_data['last_name'],
                                            organization=form.cleaned_data['organization'])
            user.is_superuser = False
            user.is_supervisor = False

            # User role checkboxes
            if 'user_type' in request.POST:
                roles = request.POST.getlist('user_type')
                user.is_superuser = 'admin' in roles
                user.is_supervisor = 'supervisor' in roles
                user.is_reentry_coordinator = 'reentry_coordinator' in roles
                user.is_community_outreach_worker = 'community_outreach_worker' in roles
                user.is_service_provider = 'service_provider' in roles
                user.is_resource_coordinator = 'resource_coordinator' in roles
            else:
                form.clean_roles()
            
            user.save()
            
            messages.success(request, 'Added a new user to the system.')
        elif 'org_name' in request.POST:
            organization = Organization(name=request.POST['org_name'])
            organization.save()
            messages.success(request, 'Added a new organization to the system.')
    
    context['form'] = RegistrationForm()
    return render(request, 'NewEra/dashboard.html', context)

@login_required
def supervisor_dashboard(request): 
    if not request.user.is_supervisor:
        raise Http404

    admins = User.objects.filter(is_superuser=True)
    reentry_coordinators = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_reentry_coordinator=True)
    community_outreach_workers = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_community_outreach_worker=True)
    service_providers = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_service_provider=True)
    resource_coordinators = User.objects.filter(is_superuser=False).filter(is_supervisor=False).filter(is_resource_coordinator=True)
    supervisors = User.objects.filter(is_superuser=False).filter(is_supervisor=True)
    orgs = Organization.objects.all().order_by('name')
    context = { 'admins':admins, 
                'reentry_coordinators': reentry_coordinators, 
                'community_outreach_workers': community_outreach_workers, 
                'service_providers': service_providers, 
                'resource_coordinators': resource_coordinators, 
                'orgs':orgs, 'supervisors':supervisors, 'form': RegistrationForm(), 'user' : request.user}

    context['form'] = RegistrationForm()
    return render(request, 'NewEra/supervisor_dashboard.html', context)

@login_required
def edit_user(request, id):
    if not request.user.is_superuser:
        raise Http404

    user = get_object_or_404(User, id=id)

    if request.method == "POST":
        # If a user is editing themselves, get the form for only that user
        if user == request.user:
            form = EditSelfUserForm(request.POST, instance=user)
        # If an admin is editing someone else, get the appropriate form
        else:
            form = EditUserForm(request.POST, instance=user)

            if 'user_type' in request.POST:
                roles = request.POST.getlist('user_type')

                user.is_superuser = 'admin' in roles
                user.is_supervisor = 'supervisor' in roles
                user.is_reentry_coordinator = 'reentry_coordinator' in roles
                user.is_community_outreach_worker = 'community_outreach_worker' in roles
                user.is_service_provider = 'service_provider' in roles
                user.is_resource_coordinator = 'resource_coordinator' in roles

    
        if form.is_valid():
            form.save()
            user.save()

            messages.success(request, '{} successfully edited.'.format(str(user)))
            return redirect('Dashboard')
    else:
        # If a user is editing themselves, get the form for only that user
        if user == request.user:
            form = EditSelfUserForm(instance=user)
        # If an admin is editing someone else, get the appropriate form
        else:
            form = EditUserForm(instance=user)
            form.initial['user_type'] = user.get_user_types()

    return render(request, 'NewEra/edit_user.html', {'form': form, 'user': user, 'action': 'Edit'})

@login_required
def edit_org(request, id):
    if not request.user.is_superuser:
        raise Http404

    org = get_object_or_404(Organization, id=id)

    if request.method == "POST":
        
        form = EditOrganizationForm(request.POST, instance=org)
    
        if form.is_valid():

            form.save()
            org.save()

            messages.success(request, '{} successfully edited.'.format(str(org)))
            return redirect('Dashboard')
    else:
        form = EditOrganizationForm(instance=org)
    return render(request, 'NewEra/edit_org.html', {'form': form, 'org': org, 'action': 'Edit'})

@login_required
def delete_user(request, id):
    if not request.user.is_superuser:
        raise Http404

    user = get_object_or_404(User, id=id)

    if request.method == 'POST':
        # Delete the user only if they have never made a referral and there is no one on their case load
        if (user.get_referrals().count() == 0 and user.get_case_load().count() == 0):
            user.delete()
            messages.success(request, 'User successfully deleted.')
            return redirect('Dashboard')
        # Otherwise, deactivate the user
        else:
            user.is_active = False
            user.save()
            messages.success(request, '{} was made inactive.'.format(user.get_full_name()))
            return redirect('Dashboard')
    return render(request, 'NewEra/delete_user.html', {'user': user})

@login_required
def delete_org(request, id):
    if not request.user.is_superuser:
        raise Http404
        
    org = get_object_or_404(Organization, id=id)

    if request.method == 'POST':
        org.delete()
        messages.success(request, 'Organization successfully deleted.')
        return redirect('Dashboard')
    return render(request, 'NewEra/delete_org.html', {'org': org})

# endregion

# region FORMS

###################################### 
#              FORMS
######################################

@login_required
def meeting_tracker(request):
    context = {} 
    if request.user.is_superuser: 
        context['responses'] = MeetingTracker.objects.all().order_by('-date')
    elif request.user.is_supervisor:
        context['responses'] = MeetingTracker.objects.filter(user__in=User.objects.filter(organization=request.user.organization)).order_by('-date')
    elif request.user.is_superuser or request.user.is_reentry_coordinator or request.user.is_community_outreach_worker or request.user.is_service_provider or request.user.is_resource_coordinator:
        context['responses'] = MeetingTracker.objects.filter(user=request.user).order_by('-date')
    else:  
        raise Http404
    return render(request, 'NewEra/meeting_tracker.html', context)

@login_required
def create_meeting_tracker_response(request):
    context = {}
    form = MeetingTrackerForm(initial={'duration': 0})
    context['form'] = form
    context['action'] = 'Create'
    if request.method == 'POST':
        response = MeetingTracker(user=request.user)
        form = MeetingTrackerForm(request.POST, instance=response)
        
        if form.is_valid():
            response.with_who=form.cleaned_data['with_who']
            response.purpose=form.cleaned_data['purpose']
            response.neighborhood=form.cleaned_data['neighborhood']
            response.duration=form.cleaned_data['duration']
            response.notes=form.cleaned_data['notes']
            response.date=form.cleaned_data['date']
            response.time=form.cleaned_data['time']
            form.save()
            response.save()

            messages.success(request, 'Meeting Tracker Form successfully submitted!')

            return redirect('Show Meeting Tracker Response', id=response.id)
        
    else:
        response = MeetingTrackerForm()
    
    return render(request, 'NewEra/edit_meeting_tracker_response.html', context)

@login_required
def edit_meeting_tracker_response(request, id):
    
    response = MeetingTracker.objects.filter(id=id).first()
    if response.user != request.user and not(request.user.is_superuser):
        raise Http404
    context = {}
    context['action'] = 'Edit'
    
    if request.method == "POST":
        form = MeetingTrackerForm(request.POST, instance=response)
        context['form'] = form

        if form.is_valid():
            form.save()
            response.save()

            messages.success(request, 'Submission successfully edited!')
            return redirect('Show Meeting Tracker Response', id=response.id)
        
    else:
        form = MeetingTrackerForm(instance=response)
        context['form'] = form
    return render(request, 'NewEra/edit_meeting_tracker_response.html', context)

@login_required
def get_meeting_tracker_response(request, id):
    context = {}
    response = MeetingTracker.objects.filter(id=id).first()
    if response.user != request.user and not(request.user.is_superuser) and not (request.user.is_supervisor and request.user.organization == response.user.organization ):
        raise Http404
    context['response'] = response
    return render(request, 'NewEra/meeting_tracker_response.html', context)

@login_required
def delete_meeting_tracker_response(request, id):
    response = MeetingTracker.objects.filter(id=id).first()
    if response.user != request.user and not(request.user.is_superuser):
        raise Http404
    if request.method == 'POST':
        response.delete()
        messages.success(request, 'Response successfully deleted.')
        return redirect('Meeting Tracker')
    
    return render(request, 'NewEra/delete_meeting_tracker_response.html', {'response': response})

# @login_required
# def risk_assessment(request):
#     context = {} 
#     if request.user.is_superuser: 
#         context['responses'] = RiskAssessment.objects.all().order_by('-date')
#     elif request.user.is_supervisor:
#         context['responses'] = RiskAssessment.objects.filter(user__in=User.objects.filter(organization=request.user.organization)).order_by('-date')
#    elif request.user.is_superuser or request.user.is_reentry_coordinator or request.user.is_community_outreach_worker or request.user.is_service_provider or request.user.is_resource_coordinator:
#         context['responses'] = RiskAssessment.objects.filter(user=request.user).order_by('-date')
#     else:  
#         raise Http404
#     return render(request, 'NewEra/risk_assessment.html', context)

# @login_required
# def create_risk_assessment_response(request):
#     context = {}
#     form = RiskAssessmentForm()
#     context['form'] = form
#     context['action'] = 'Create'

#     if request.method == 'POST':
#         response = RiskAssessment(user=request.user)
#         form = RiskAssessmentForm(request.POST, instance=response)
        
#         if form.is_valid():
#             response.risk_level=form.cleaned_data['risk_level']
#             response.actions=form.cleaned_data['actions']
#             response.neighborhood=form.cleaned_data['neighborhood']
#             response.notes=form.cleaned_data['notes']
#             response.date=form.cleaned_data['date']
#             response.time=form.cleaned_data['time']
#             form.save()
#             response.save()

#             messages.success(request, 'Risk Assessment Form successfully submitted!')

#             if int(response.risk_level) > 2:
#                 response.send_email()

#             return redirect('Show Risk Assessment Response', id=response.id)
        
#     else:
#         response = RiskAssessmentForm()
    
#     return render(request, 'NewEra/edit_risk_assessment_response.html', context)

# @login_required
# def edit_risk_assessment_response(request, id):
#     response = RiskAssessment.objects.filter(id=id).first()
#     if response.user != request.user and not(request.user.is_superuser):
#         raise Http404
#     context = {}
#     context['action'] = 'Edit'

#     if request.method == "POST":
#         form = RiskAssessmentForm(request.POST, instance=response)
#         context['form'] = form

#         if form.is_valid():
#             form.save()
#             response.save()

#             messages.success(request, 'Submission successfully edited!')
#             return redirect('Show Risk Assessment Response', id=response.id)
        
#     else:
#         form = RiskAssessmentForm(instance=response)
#         context['form'] = form
#     return render(request, 'NewEra/edit_risk_assessment_response.html', context)

# @login_required
# def get_risk_assessment_response(request, id):
#     context = {}
#     response = RiskAssessment.objects.filter(id=id).first()
#     if response.user != request.user and not(request.user.is_superuser) and not (request.user.is_supervisor and request.user.organization == response.user.organization ):
#         raise Http404
#     context['response'] = response
#     return render(request, 'NewEra/risk_assessment_response.html', context)

# @login_required
# def delete_risk_assessment_response(request, id):
#     response = RiskAssessment.objects.filter(id=id).first()
#     if response.user != request.user and not(request.user.is_superuser):
#         raise Http404
#     if request.method == 'POST':
#         response.delete()
#         messages.success(request, 'Response successfully deleted.')
#         return redirect('Risk Assessment')
    
#     return render(request, 'NewEra/delete_risk_assessment_response.html', {'response': response})

# @login_required
# def biweekly(request):
#     context = {} 
#     if request.user.is_superuser: 
#         context['responses'] = Biweekly.objects.all().order_by('-date')
#     elif request.user.is_supervisor:
#         context['responses'] = Biweekly.objects.filter(user__in=User.objects.filter(organization=request.user.organization)).order_by('-date')
#    elif request.user.is_superuser or request.user.is_reentry_coordinator or request.user.is_community_outreach_worker or request.user.is_service_provider or request.user.is_resource_coordinator:
#         context['responses'] = Biweekly.objects.filter(user=request.user).order_by('-date')
#     else:  
#         raise Http404
#     return render(request, 'NewEra/biweekly.html', context)

# @login_required
# def create_biweekly_response(request):
#     context = {}
#     cases = CaseLoadUser.objects.filter(user=request.user)
#     hours = 0
#     for case in cases:
#       notes = Note.objects.filter(case=case, date__gte=timezone.now()-timedelta(days=14)).exclude(activity_type="Group Meeting")
#       for note in notes:
#           hours += note.hours

#     form = BiweeklyForm(initial={'work_hours': hours})

#     context['form'] = form
#     context['action'] = 'Create'
#     if request.method == 'POST':
#         response = Biweekly(user=request.user)
#         form = BiweeklyForm(request.POST, instance=response)
        
#         if form.is_valid():
#             response.pay_start=form.cleaned_data['pay_start']
#             response.pay_end=form.cleaned_data['pay_end']
#             response.work_hours=form.cleaned_data['work_hours']
#             response.special_hours=form.cleaned_data['special_hours']
#             response.meeting_hours=form.cleaned_data['meeting_hours']
#             response.num_meetings=form.cleaned_data['num_meetings']
#             response.other=form.cleaned_data['other']
#             response.num_served=form.cleaned_data['num_served']
#             response.location=form.cleaned_data['location']
#             response.intervention=form.cleaned_data['intervention']
#             response.num_response=form.cleaned_data['num_responses']
#             response.num_events=form.cleaned_data['num_events']
#             response.num_visits=form.cleaned_data['num_visits']
#             response.num_mediations=form.cleaned_data['num_mediations']
#             response.first_summary=form.cleaned_data['first_summary']
#             response.second_summary=form.cleaned_data['second_summary']
#             response.date=form.cleaned_data['date']
#             form.save()
#             response.save()

#             messages.success(request, 'Bi-Weekly Attendance Form successfully submitted!')

#             return redirect('Show Biweekly Response', id=response.id)
        
#     else:
#         response = BiweeklyForm()
    
#     return render(request, 'NewEra/edit_biweekly_response.html', context)

# @login_required
# def edit_biweekly_response(request, id):
#     response = Biweekly.objects.filter(id=id).first()
#     if response.user != request.user and not(request.user.is_superuser):
#         raise Http404
#     context = {}
#     context['action'] = 'Edit'

#     if request.method == "POST":
#         form = BiweeklyForm(request.POST, instance=response)
#         context['form'] = form

#         if form.is_valid():
#             form.save()
#             response.save()

#             messages.success(request, 'Submission successfully edited!')
#             return redirect('Show Biweekly Response', id=response.id)
        
#     else:
#         form = BiweeklyForm(instance=response)
#         context['form'] = form
#     return render(request, 'NewEra/edit_biweekly_response.html', context)

# @login_required
# def get_biweekly_response(request, id):
#     context = {}
#     response = Biweekly.objects.filter(id=id).first()
#     if response.user != request.user and not(request.user.is_superuser) and not (request.user.is_supervisor and request.user.organization == response.user.organization ):
#         raise Http404
#     context['response'] = response
#     return render(request, 'NewEra/biweekly_response.html', context)

# @login_required
# def delete_biweekly_response(request, id):
#     response = Biweekly.objects.filter(id=id).first()
#     if response.user != request.user and not(request.user.is_superuser):
#         raise Http404
#     if request.method == 'POST':
#         response.delete()
#         messages.success(request, 'Response successfully deleted.')
#         return redirect('Biweekly')
    
#     return render(request, 'NewEra/delete_biweekly_response.html', {'response': response})

# endregion

# region TAGS

###################################### 
#              TAGS 
######################################

@login_required
def tags(request):
    if not request.user.is_superuser:
        raise Http404
    context = {
        'tags': Tag.objects.all()
    }
    return render(request, 'NewEra/tags.html', context)

@login_required
def create_tag(request):
    if not request.user.is_superuser:
        raise Http404
    context = {}
    form = TagForm()
    context['form'] = form
    context['action'] = 'Create'

    if request.method == 'POST':
        tag = Tag()
        form = TagForm(request.POST, instance=tag)
        
        if form.is_valid():
            tag.tag_type=form.cleaned_data['tag_type']
            form.save()
            tag.save()

            messages.success(request, 'Tag successfully created!')

            return redirect('Tags')
    else:
        tag = TagForm()

    return render(request, 'NewEra/edit_tag.html', context)

@login_required
def edit_tag(request, id):
    if not request.user.is_superuser:
        raise Http404
    tag = get_object_or_404(Tag, id=id)

    if request.method == "POST":
        form = TagForm(request.POST, instance=tag)
    
        if form.is_valid():
            form.save()
            tag.save()

            messages.success(request, '{} successfully edited.'.format(tag.name))
            return redirect('Tags')
    else:
        form = TagForm(instance=tag)
    return render(request, 'NewEra/edit_tag.html', {'form': form, 'tag': tag, 'action': 'Edit'})

@login_required
def delete_tag(request, id):
    if not request.user.is_superuser:
        raise Http404
    tag = get_object_or_404(Tag, id=id)

    if request.method == 'POST':
        tag.delete()
        messages.success(request, '{} successfully deleted.'.format(tag.name))
        return redirect('Tags')

    return render(request, 'NewEra/delete_tag.html', {'tag': tag})

# endregion

# region KPI SPREADSHEET EXPORT

###################################### 
# KPI SPREADSHEET EXPORT 
######################################

@login_required
def select_data(request):
    if not request.user.is_superuser:
        raise Http404
    form = SelectDataTimeframe()
    return render(request, 'NewEra/select_data.html', {'form': form})

def select_referral_data(request):
    if not request.user.is_superuser:
        raise Http404
    form = SelectDataTimeframe()
    return render(request, 'NewEra/select_referral_data.html', {'form': form})

@login_required 
def export_selected_data(request):
    if not request.user.is_superuser:
        raise Http404
    if request.method == "GET":
        form = SelectDataTimeframe()
        return render(request, 'NewEra/select_data.html', {'form': form})
    else:
        form  =  SelectDataTimeframe(request.POST)
        if form.is_valid():
            # risk_assessments = RiskAssessment.objects.all().filter(date__range=(form.cleaned_data["start_date"],form.cleaned_data["end_date"]))
            meeting_trackers  = MeetingTracker.objects.all().filter(date__range=(form.cleaned_data["start_date"],form.cleaned_data["end_date"])).order_by('user')

            # Define the workbook and sheet
            wb = Workbook()
            ws = wb.active

            # Set the bold font
            bold = Font(bold=True)

            # # Create Header row
            # ws['A1'].font = bold
            # ws['B1'].font = bold
            # ws['C1'].font = bold
            # ws['D1'].font = bold
            # ws['E1'].font = bold
            # ws['F1'].font = bold

            # ws['A1'] = "Date"
            # ws['B1'] = "User"
            # ws['C1'] = "Neighborhood"
            # # ws['D1'] = "Risk Level"
            # ws['D1'] = "Actions"
            # ws['E1'] = "Notes"

            # for r in risk_assessments:
            #     date = r.date.__str__()
            #     user = r.user.__str__()
            #     neighborhood = r.neighborhood
            #     risk_level = r.risk_level
            #     actions = r.actions
            #     notes = r.notes
            #     ws.append([date,user, neighborhood, risk_level, actions, notes])

            

            # ws = wb.worksheets[0]
            # ws.auto_filter.ref = ws.dimensions
            
            # ws.title = "Risk Level Assessments"
            # ws.column_dimensions[get_column_letter(1)].width = 50
            # ws.column_dimensions[get_column_letter(2)].width = 50
            # ws.column_dimensions[get_column_letter(3)].width = 50
            # ws.column_dimensions[get_column_letter(4)].width = 50
            # ws.column_dimensions[get_column_letter(5)].width = 50
            # ws.column_dimensions[get_column_letter(6)].width = 50

            ws1 = wb.create_sheet("Meeting Tracker Forms")
            ws = ws1

            # Create Header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold
            ws['E1'].font = bold
            ws['F1'].font = bold
            ws['G1'].font = bold

            ws['A1'] = "Date"
            ws['B1'] = "User"
            ws['C1'] = "With Who"
            ws['D1'] = "Purpose"
            ws['E1'] = "Neighborhood"
            ws['F1'] = "Duration"
            ws['G1'] = "Notes"

            for m in meeting_trackers:
                date = m.date.__str__()
                user = m.user.__str__()
                with_who = m.with_who
                purpose = m.purpose
                neighborhood = m.neighborhood
                duration = m.duration
                notes = m.notes
                ws.append([date, user, with_who, purpose, neighborhood, duration,notes ])

            ws.column_dimensions[get_column_letter(1)].width = 50
            ws.column_dimensions[get_column_letter(2)].width = 50
            ws.column_dimensions[get_column_letter(3)].width = 50
            ws.column_dimensions[get_column_letter(4)].width = 50
            ws.column_dimensions[get_column_letter(5)].width = 50
            ws.column_dimensions[get_column_letter(6)].width = 50
            ws.column_dimensions[get_column_letter(7)].width = 50

            ws.auto_filter.ref = ws.dimensions

            # ws2 = wb.create_sheet("Average Risk Level")
            # ws = ws2

            # # Create Header row
            # ws['A1'].font = bold
            # ws['B1'].font = bold

            # ws['A1'] = "Neighborhood"
            # ws['B1'] = "Average Risk Level"

            # risk_dict = dict()
            # for r in risk_assessments:
            #     user = r.user.__str__()
            #     neighborhood = r.neighborhood
            #     risk_level = int(r.risk_level)
            #     if neighborhood not in risk_dict:
            #         risk_dict[neighborhood] = []
            #     else:
            #         risk_dict[neighborhood].append(risk_level)

            # new_list = []

            # for key in risk_dict:
            #     lst = risk_dict[key]
            #     if len(lst) > 0:
            #         new_list.append( (key, round( (sum(lst) / len(lst) ), 2) ) )
            #     else:
            #         new_list.append((key, 0))

            # new_list.sort(key = lambda x: x[1], reverse = True)
            
            # for (k,v) in new_list:
            #     ws.append([k,v])

            # ws.column_dimensions[get_column_letter(1)].width = 50
            # ws.column_dimensions[get_column_letter(2)].width = 50

            # ws2 = wb.create_sheet("Actions Count")
            # ws = ws2

            # # Create Header row
            # ws['A1'].font = bold
            # ws['B1'].font = bold

            # ws['A1'] = "Action"
            # ws['B1'] = "Count"

            # risk_dict = dict()
            # for r in risk_assessments:
            #     user = r.user.__str__()
            #     neighborhood = r.neighborhood
            #     actions = r.actions
            #     if actions not in risk_dict:
            #         risk_dict[actions] = 0
            #     else:
            #         risk_dict[actions]+=1
            
            # new_list = []
            # for key in risk_dict:
            #     new_list.append( (key, risk_dict[key]) )
            
            # new_list.sort(key = lambda x: x[1], reverse = True)
            
            # for (k,v) in new_list:
            #     ws.append([k,v])

            # ws.column_dimensions[get_column_letter(1)].width = 50
            # ws.column_dimensions[get_column_letter(2)].width = 50
            

            ws3 = wb.create_sheet("Meeting With Who Count")
            ws = ws3

            # Create Header row
            ws['A1'].font = bold
            ws['B1'].font = bold

            ws['A1'] = "Meeting With Who"
            ws['B1'] = "Count"

            risk_dict = dict()
            for m in meeting_trackers:
                with_who = m.with_who
                if with_who not in risk_dict:
                    risk_dict[with_who] = 0
                else:
                    risk_dict[with_who]+=1
            
            new_list = []
            for key in risk_dict:
                new_list.append( (key, risk_dict[key]) )
            
            new_list.sort(key = lambda x: x[1], reverse = True)
            
            for (k,v) in new_list:
                ws.append([k,v])
                

            ws.column_dimensions[get_column_letter(1)].width = 50
            ws.column_dimensions[get_column_letter(2)].width = 50

            ws4 = wb.create_sheet("Outreach Worker Time")
            ws = ws4

            # Create Header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold

            ws['A1'] = "Outreach Worker"
            ws['B1'] = "With Who"
            ws['C1'] = "Total Duration (Hours)"
            new_dict = dict()
            for m in meeting_trackers:
                user = m.user.__str__()
                with_who = m.with_who
                durations = m.duration
                if (user, purpose) not in risk_dict:
                    new_dict[(user, with_who)] = durations
                else:
                    new_dict[(user, with_who)] += durations
            
            for (user, with_who) in new_dict:
                ws.append([user, with_who, new_dict[(user, with_who)]])

            ws.column_dimensions[get_column_letter(1)].width = 50
            ws.column_dimensions[get_column_letter(2)].width = 50
            ws.column_dimensions[get_column_letter(3)].width = 50
            ws.auto_filter.ref = ws.dimensions

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = "attachment; filename=newera412_selected_data_spreadsheet.xlsx"
            wb.save(response)

            return response
        else:
            return Http404

# Export data on resources and referrals
@login_required
def export_data(request):
    if not request.user.is_superuser:
        raise Http404
    if request.method == "GET":
        form = SelectDataTimeframe()
        return render(request, 'NewEra/select_referral_data.html', {'form': form})
    else:
        form  =  SelectDataTimeframe(request.POST)
        if form.is_valid():
            # Get resources
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]
            resources = Resource.objects.filter(is_active=True, referrals__referral_date__gte=start_date, referrals__referral_date__lte=end_date).distinct()
            # student_resources = Resource.objects.filter(is_active=True, student_referrals__referral_date__gte=start_date, student_referrals__referral_date__lte=end_date).distinct()
    
            # Define the workbook and sheet
            wb = Workbook()
            ws = wb.active

            # Set the bold font
            bold = Font(bold=True)

            # Create Header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold

            ws['A1'] = "Resource"
            ws['B1'] = "# Referrals"
            ws['C1'] = "# Accessed Referrals"
            ws['D1'] = "# Views"

            # Write resources
            for r in resources:
                # Get name
                name = r.name
                # Get referrals including the resource
                referrals_count = r.referrals.count()
                # Get accessed referrals
                accessed_count = r.referrals.exclude(date_accessed=None).count()
                # Get clicks
                clicks = r.clicks
                # Write to the Excel file
                ws.append([name, referrals_count, accessed_count, clicks])
            
            ws = wb.worksheets[0]
            ws.title = "General Referral Resources"
            ws.column_dimensions[get_column_letter(1)].width = 60
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 20

            # Export breakdown by tag

            # ws_student = wb.create_sheet("Student Referral Resources")
            # ws = ws_student

             # Create Header row
            # ws['A1'].font = bold
            # ws['B1'].font = bold
            # ws['C1'].font = bold
            # ws['D1'].font = bold

            # ws['A1'] = "Resource"
            # ws['B1'] = "# Referrals"
            # ws['C1'] = "# Accessed Referrals"
            # ws['D1'] = "# Views"

            # # Write resources
            # for r in student_resources:
            #     # Get name
            #     name = r.name
            #     # Get referrals including the resource
            #     referrals_count = r.student_referrals.count()
            #     # Get accessed referrals
            #     accessed_count = r.student_referrals.exclude(date_accessed=None).count()
            #     # Get clicks
            #     clicks = r.clicks
            #     # Write to the Excel file
            #     ws.append([name, referrals_count, accessed_count, clicks])
            
            # ws.column_dimensions[get_column_letter(1)].width = 60
            # ws.column_dimensions[get_column_letter(2)].width = 20
            # ws.column_dimensions[get_column_letter(3)].width = 20
            # ws.column_dimensions[get_column_letter(4)].width = 20

            # Export breakdown by tag

            ws1 = wb.create_sheet("By Tag")
            ws = ws1

            # Create Header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold
            ws['E1'].font = bold
            # ws['F1'].font = bold
            # ws['G1'].font = bold

            ws['A1'] = "Tag"
            ws['B1'] = "# Resources"
            ws['C1'] = "# General Referrals"
            ws['D1'] = "# Accessed General Referrals"
            ws['E1'] = "# Views"

            # Get tags
            tags = Tag.objects.all().exclude(resource=None)

            for t in tags:
                # Get name
                name = t.name

                resources_count = 0

                general_referrals_count_by_tag = 0
                # student_referrals_count_by_tag = 0
                accessed_count = 0
                # student_accessed_count = 0
                clicks = 0

                for r in resources:
                    if t in r.tags.all():
                        # Increment number of resources with this tag
                        resources_count += 1
                        # Get general referrals associated with the tag
                        general_referrals_count_by_tag += r.referrals.count()
                        # Get student referrals associated with the tag
                        # student_referrals_count_by_tag += r.student_referrals.count()
                        # Get accessed referrals by tag
                        accessed_count += r.referrals.exclude(date_accessed=None).count()
                        # Get accessed student referrals by tag
                        # student_accessed_count += r.student_referrals.exclude(date_accessed=None).count()
                        # Get clicks
                        clicks += r.clicks

                # Write to the Excel file
                ws.append([name, resources_count, general_referrals_count_by_tag, accessed_count, clicks])
            
            ws.column_dimensions[get_column_letter(1)].width = 30
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 30
            ws.column_dimensions[get_column_letter(5)].width = 20
            # ws.column_dimensions[get_column_letter(6)].width = 30
            # ws.column_dimensions[get_column_letter(7)].width = 20


            # Export user data

            ws2 = wb.create_sheet("By User")
            ws = ws2

            # Create Header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold
            ws['E1'].font = bold
            # ws['F1'].font = bold
            # ws['G1'].font = bold

            ws['A1'] = "User"
            ws['B1'] = "# Case Load Users"
            ws['C1'] = "# General Referrals"
            ws['D1'] = "# Accessed General Referrals"
            # ws['E1'] = "# Student Referrals"
            # ws['F1'] = "# Accessed Student Referrals"
            # ws['G1'] = "Date of Last Referral"
            ws['E1'] = "Date of Last Referral"

            # Get users
            users = User.objects.all()

            for u in users:
                case_load_count = u.get_case_load().count()
                referrals_count = u.get_referrals().count()
                accessed_referrals_count = u.get_referrals().exclude(date_accessed=None).count()
                # student_referrals_count = u.get_student_referrals().count()
                # accessed_student_referrals_count = u.get_student_referrals().exclude(date_accessed=None).count()
                last_general_referral_date = u.get_referrals().order_by('-referral_date').first()
                # last_student_referral_date = u.get_referrals().order_by('-referral_date').first()
                if last_general_referral_date is not None:
                    last_referral_date = last_general_referral_date.referral_date.strftime('%m-%d-%Y')
                else:
                    last_referral_date = "No referrals made"

                # Write to the Excel file
                ws.append([str(u), case_load_count, referrals_count, accessed_referrals_count, last_referral_date])

            ws.column_dimensions[get_column_letter(1)].width = 30
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 30
            ws.column_dimensions[get_column_letter(5)].width = 20
            # ws.column_dimensions[get_column_letter(6)].width = 30
            # ws.column_dimensions[get_column_letter(7)].width = 20


            # Export referral data by phone
            ws3 = wb.create_sheet("By Referred Phone")
            ws = ws3

            # Create header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold
            ws['E1'].font = bold

            ws['A1'] = "Phone"
            ws['B1'] = "Case Load User"
            ws['C1'] = "# Referrals"
            ws['D1'] = "# Accessed Referrals"
            ws['E1'] = "Date of Last Referral"

            # Get referrals
            referrals = Referral.objects.filter(referral_date__gte=start_date, referral_date__lte=end_date).order_by('-referral_date')

            # Sets to keep track of phones and emails seen
            phones = set()
            emails = set()
            case_load_dict = dict()
            referrals_dict = dict()
            accessed_referrals_dict = dict()
            last_referral_dict = dict()

            for r in referrals:
                # Organize and count by phone
                if r.phone:
                    # Format phone number
                    if (len(r.phone) == 10):
                        phone_number = "(" + r.phone[0:3] + ") " + r.phone[3:6] + "-" + r.phone[6:10]
                    else:
                        phone_number = r.phone[0] + " (" + r.phone[1:4] + ") " + r.phone[4:7] + "-" + r.phone[7:11]
                    referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict = export_attribute(phone_number, phones, referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict, r)
                # Organize and count by email
                if r.email:
                    referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict = export_attribute(r.email, emails, referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict, r)

            for p in phones:
                # Write to the Excel file
                ws.append([p, case_load_dict[p], referrals_dict[p], accessed_referrals_dict[p], last_referral_dict[p]])

            ws.column_dimensions[get_column_letter(1)].width = 30
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 20


            # Export referral data by email
            ws4 = wb.create_sheet("By Referred Email")
            ws = ws4

            # Create header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold
            ws['E1'].font = bold

            ws['A1'] = "Email"
            ws['B1'] = "Case Load User"
            ws['C1'] = "# Referrals"
            ws['D1'] = "# Accessed Referrals"
            ws['E1'] = "Date of Last Referral"

            for e in emails:
                # Write to the Excel file
                ws.append([e, case_load_dict[e], referrals_dict[e], accessed_referrals_dict[e], last_referral_dict[e]])

            ws.column_dimensions[get_column_letter(1)].width = 30
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 20

            # Export referral data by student
            # ws_stud = wb.create_sheet("By Student")
            # ws = ws_stud

            # Create header row
            # ws['A1'].font = bold
            # ws['B1'].font = bold
            # ws['C1'].font = bold
            # ws['D1'].font = bold
            # ws['E1'].font = bold
            # ws['F1'].font = bold
            # ws['G1'].font = bold
            # ws['H1'].font = bold

            # ws['A1'] = "Student"
            # ws['B1'] = "Email"
            # ws['C1'] = "Phone"
            # ws['D1'] = "Referrer"
            # ws['E1'] = "School"
            # ws['F1'] = "# Referrals"
            # ws['G1'] = "# Accessed Referrals"
            # ws['H1'] = "Date of Last Referral"

            # Get referrals
            # referrals = StudentReferral.objects.filter(referral_date__gte=start_date, referral_date__lte=end_date).order_by('-referral_date')

            # Sets to keep track of phones and emails seen
            # students = set()
            # schools_dict = dict()
            # referrer_dict = dict()
            # phones_dict = dict()
            # emails_dict = dict()
            # referrals_dict = dict()
            # accessed_referrals_dict = dict()
            # last_referral_dict = dict()

            # for r in referrals:
            #     # Format phone number
            #     # if (len(r.phone) == 10):
            #     #     phone_number = "(" + r.phone[0:3] + ") " + r.phone[3:6] + "-" + r.phone[6:10]
            #     # else:
            #     #     phone_number = r.phone[0] + " (" + r.phone[1:4] + ") " + r.phone[4:7] + "-" + r.phone[7:11]
            #     referrals_dict, accessed_referrals_dict, referrer_dict, schools_dict, phones_dict, emails_dict, last_referral_dict = export_student_attribute(r.student, students, referrals_dict, accessed_referrals_dict, referrer_dict, schools_dict, phones_dict, emails_dict, last_referral_dict, r)

            # for s in students:
            #     # Write to the Excel file
            #     ws.append([s.get_full_name(), emails_dict[s], phones_dict[s], referrer_dict[s], schools_dict[s], referrals_dict[s], accessed_referrals_dict[s], last_referral_dict[s]])

            # ws.column_dimensions[get_column_letter(1)].width = 25
            # ws.column_dimensions[get_column_letter(2)].width = 30
            # ws.column_dimensions[get_column_letter(3)].width = 20
            # ws.column_dimensions[get_column_letter(4)].width = 20
            # ws.column_dimensions[get_column_letter(5)].width = 20
            # ws.column_dimensions[get_column_letter(6)].width = 20
            # ws.column_dimensions[get_column_letter(7)].width = 20

            # Export referral data by team
            ws5 = wb.create_sheet("By Team")
            ws = ws5

            # Set to keep track of teams seen
            teams = set()
            team_cases = dict()
            team_referrals = dict()
            # team_student_referrals = dict()
            team_accessed_referrals = dict()
            # team_accessed_student_referrals = dict()
            team_last_referral_date = dict()
            users = User.objects.all()

            for u in users:

                teams.add(u.team)

                if u.team in team_cases.keys():
                    team_cases[u.team] += u.get_case_load().count()
                else:
                    team_cases[u.team] = u.get_case_load().count()
                
                if u.team in team_referrals.keys():
                    team_referrals[u.team] += u.get_referrals().filter(referral_date__gte=start_date, referral_date__lte=end_date).count()
                else:
                    team_referrals[u.team] = u.get_referrals().filter(referral_date__gte=start_date, referral_date__lte=end_date).count()

                # if u.team in team_student_referrals.keys():
                #     team_student_referrals[u.team] += u.get_student_referrals().filter(referral_date__gte=start_date, referral_date__lte=end_date).count()
                # else:
                #     team_student_referrals[u.team] = u.get_student_referrals().filter(referral_date__gte=start_date, referral_date__lte=end_date).count()
                
                if u.team in team_accessed_referrals.keys():
                    team_accessed_referrals[u.team] += u.get_referrals().exclude(date_accessed=None).filter(referral_date__gte=start_date, referral_date__lte=end_date).count()
                else:
                    team_accessed_referrals[u.team] = u.get_referrals().exclude(date_accessed=None).filter(referral_date__gte=start_date, referral_date__lte=end_date).count()

                # if u.team in team_accessed_student_referrals.keys():
                #     team_accessed_student_referrals[u.team] += u.get_student_referrals().exclude(date_accessed=None).filter(referral_date__gte=start_date, referral_date__lte=end_date).count()
                # else:
                #     team_accessed_student_referrals[u.team] = u.get_student_referrals().exclude(date_accessed=None).filter(referral_date__gte=start_date, referral_date__lte=end_date).count()
                
                last_referral_date = u.get_referrals().order_by('-referral_date').first()
                # last_student_referral_date = u.get_student_referrals().order_by('-referral_date').first()

                if u.team in team_last_referral_date:
                    if last_referral_date:
                        team_last_referral_date[u.team] = last_referral_date
                    else:
                        team_last_referral_date[u.team] = "No referrals made"
                else:
                    if last_referral_date:
                        team_last_referral_date[u.team] = last_referral_date
                    else:
                        team_last_referral_date[u.team] = "No referrals made"

            for key in team_last_referral_date:
                if team_last_referral_date[key] != "No referrals made":
                    team_last_referral_date[key] = team_last_referral_date[key].referral_date.strftime('%m-%d-%Y')

            # Create header row
            ws['A1'].font = bold
            ws['B1'].font = bold
            ws['C1'].font = bold
            ws['D1'].font = bold
            ws['E1'].font = bold
            # ws['F1'].font = bold
            # ws['G1'].font = bold

            ws['A1'] = "Team"
            ws['B1'] = "# Case Load Users"
            ws['C1'] = "# General Referrals"
            ws['D1'] = "# Accessed General Referrals"
            # ws['E1'] = "# Student Referrals"
            # ws['F1'] = "# Accessed Student Referrals"
            ws['E1'] = "Date of Last Referral"

            for t in teams:
                # Write to the Excel file
                ws.append([t, team_cases[t], team_referrals[t], team_accessed_referrals[t], team_last_referral_date[t]])

            ws.column_dimensions[get_column_letter(1)].width = 30
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 30
            ws.column_dimensions[get_column_letter(4)].width = 30
            ws.column_dimensions[get_column_letter(5)].width = 30
            # ws.column_dimensions[get_column_letter(6)].width = 30
            # ws.column_dimensions[get_column_letter(7)].width = 25

            # Save and download the Excel file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f"attachment; filename=newera412_referral_data_{start_date}_to_{end_date}.xlsx"
            wb.save(response)

            return response
        else:
            return Http404

# Helper method to export attributes in a sheet
def export_attribute(attr, attr_set, referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict, r):
    # Add email to set
    if (attr not in attr_set):
        attr_set.add(attr)
        referrals_dict[attr] = 0
        accessed_referrals_dict[attr] = 0

    # Add the case load user to the dictionary
    if (r.caseUser):
        case_load_dict[attr] = r.caseUser.get_full_name()
    else:
        case_load_dict[attr] = "-"

    # Add referrals given to this phone number
    referrals_dict[attr] += 1

    if (r.date_accessed):
        # Add accessed referrals
        accessed_referrals_dict[attr] += 1

    # Get the last referral made
    last_referral_dict[attr] = r.referral_date.strftime('%m-%d-%Y')

    return referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict